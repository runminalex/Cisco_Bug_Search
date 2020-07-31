import openpyxl
from openpyxl.styles import Font

def ciscopsirt(dict_source, name ='vulnerability'):
    # Create the workbook and sheet for Excel
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    critical = Font(bold='True', color='00FF0000')
    high = Font(color='00FF6600')
    # titles = ['advisoryId', 'advisoryTitle', 'bugIDs', 'cves', 'cvssBaseScore', \
    #          'cwe', 'firstPublished', 'lastUpdated', 'productNames', 'publicationUrl', 'sir', \
    #          'summary']
    titles = ['advisoryId', 'advisoryTitle', 'Severity' , 'bugIDs', 'cves', 'cvssBaseScore', \
             'cwe', 'firstPublished', 'productNames' , 'publicationUrl'\
             ]
    # openpyxl does things based on 1 instead of 0
    row = 1
    column = 1
    for title in titles:
        sheet.cell(row=row, column=column, value=title).font = Font(bold='True')
        column += 1
    dicts = dict_source['advisories']
    for source in dicts:
        column = 1
        row += 1
        # for keys, values in dicts.items():
        for title in titles:
            # Put the key in the first column for each key in the dictionary

            if title == 'bugIDs' or title == 'cves' or title == 'cwe' or title == 'productNames' :
                data = source[title][0]
                sheet.cell(row=row, column=column, value=data)
            elif title == 'advisoryId' or title == 'advisoryTitle' or title == 'cvssBaseScore' or title == 'firstPublished' or title == 'lastUpdated' or title == 'publicationUrl' or title == 'summary':
                sheet.cell(row=row, column=column, value=source[title])
            elif title == 'Severity':
                if source['sir'] == 'Critical':
                    sheet.cell(row=row, column=column, value=source['sir']).font = critical
                    # highlighter.fill = openpyxl.font(b=True, color='FF0000')
                elif source['sir'] == 'High':
                    sheet.cell(row=row, column=column, value=source['sir']).font = high
                else:
                    sheet.cell(row=row, column=column, value=source['sir'])
            else:
                break

            column += 1
            # for element in values:
            # # Put the element in each adjacent column for each element in the tuple
            # sheet.cell(row=row, column=column, value=element)
            # column += 1
            # row += 1
    for title in titles:
        if title == 'publicationUrl':
            sheet.column_dimensions['J'].width = 100
        elif title == 'bugIDs':
            sheet.column_dimensions['D'].width = 11
        elif title == 'productNames':
            sheet.column_dimensions['I'].width = 42
        elif title == 'firstPublished':
            sheet.column_dimensions['H'].width = 18
        elif title == 'cves':
            sheet.column_dimensions['E'].width = 14.3
        elif title == 'advisoryId':
            sheet.column_dimensions['A'].width = 26
        elif title == 'advisoryTitle':
            sheet.column_dimensions['B'].width = 85
        elif title == 'cvssBaseScore':
            sheet.column_dimensions['F'].width = 12.3
        elif title == 'cwe':
            sheet.column_dimensions['G'].width = 8.3
    workbook.save(filename=name + '.xlsx')



if __name__ == '__main__':
    # dict = {'advisoryId': 'cisco-sa-sbswitch-session-JZAS5jnY', 'advisoryTitle': 'Cisco Small Business Smart and Managed Switches Session Management Vulnerability'}
    dict = {'advisories': [{'advisoryId': 'cisco-sa-sbswitch-session-JZAS5jnY', 'advisoryTitle': 'Cisco Small Business Smart and Managed Switches Session Management Vulnerability', 'bugIDs': ['CSCvt53117'], 'ipsSignatures': ['NA'], 'cves': ['CVE-2020-3297'], 'cvrfUrl': 'https://tools.cisco.com/security/center/contentxml/CiscoSecurityAdvisory/cisco-sa-sbswitch-session-JZAS5jnY/cvrf/cisco-sa-sbswitch-session-JZAS5jnY_cvrf.xml', 'cvssBaseScore': '8.1', 'cwe': ['CWE-287'], 'firstPublished': '2020-07-01T16:00:00', 'lastUpdated': '2020-07-01T16:00:00', 'productNames': ['Cisco Small Business 200 Series Smart Switches ', 'Cisco Small Business 250 Series Smart Switches Software '], 'publicationUrl': 'https://tools.cisco.com/security/center/content/CiscoSecurityAdvisory/cisco-sa-sbswitch-session-JZAS5jnY', 'sir': 'High', 'summary': '\n<p>A vulnerability in session management for the web-based interface of Cisco Small Business Smart and Managed Switches could allow an unauthenticated, remote attacker to defeat authentication protections and gain unauthorized access to the management interface. The attacker could obtain the privileges of the highjacked session account, which could include <em>administrator </em>privileges on the device.</p>\n<p>The vulnerability is due to the use of weak entropy generation for session identifier values. An attacker could exploit this vulnerability to determine a current session identifier through brute force and reuse that session identifier to take over an ongoing session. In this way, an attacker could take actions within the management interface with privileges up to the level of the administrative user.</p>\n<p>Cisco has released software updates that address this vulnerability for devices that have not reached the end of software maintenance. There are no workarounds that address this vulnerability.</p>\n<p>This advisory is available at the following link:<br><a href="https://tools.cisco.com/security/center/content/CiscoSecurityAdvisory/cisco-sa-sbswitch-session-JZAS5jnY" target="_blank">https://tools.cisco.com/security/center/content/CiscoSecurityAdvisory/cisco-sa-sbswitch-session-JZAS5jnY</a></p>\n'}, {'advisoryId': 'cisco-sa-dnac-info-disc-6xsCyDYy', 'advisoryTitle': 'Cisco Digital Network Architecture Center Information Disclosure Vulnerability', 'bugIDs': ['CSCvn19092'], 'ipsSignatures': ['NA'], 'cves': ['CVE-2020-3391'], 'cvrfUrl': 'https://tools.cisco.com/security/center/contentxml/CiscoSecurityAdvisory/cisco-sa-dnac-info-disc-6xsCyDYy/cvrf/cisco-sa-dnac-info-disc-6xsCyDYy_cvrf.xml', 'cvssBaseScore': '6.5', 'cwe': ['CWE-200'], 'firstPublished': '2020-07-01T16:00:00', 'lastUpdated': '2020-07-01T16:00:00', 'productNames': ['Cisco Digital Network Architecture Center (DNA Center) '], 'publicationUrl': 'https://tools.cisco.com/security/center/content/CiscoSecurityAdvisory/cisco-sa-dnac-info-disc-6xsCyDYy', 'sir': 'Medium', 'summary': '\r\n<p>A vulnerability in Cisco&nbsp;Digital Network Architecture (DNA) Center could allow an authenticated, remote attacker to view sensitive information in clear text.</p>\r\n<p>The vulnerability is due to insecure storage of certain unencrypted credentials on an affected device. An attacker could exploit this vulnerability by viewing the network device configuration and obtaining credentials that they may not normally have access to. A successful exploit could allow the attacker to use those credentials to discover and manage network devices.</p>\r\n<p>Cisco has released software updates that address this vulnerability. There are no workarounds that address this vulnerability.</p>\r\n<p>This advisory is available at the following link:<br><a href="https://tools.cisco.com/security/center/content/CiscoSecurityAdvisory/cisco-sa-dnac-info-disc-6xsCyDYy" target="_blank">https://tools.cisco.com/security/center/content/CiscoSecurityAdvisory/cisco-sa-dnac-info-disc-6xsCyDYy</a></p>\r\n'}, {'advisoryId': 'cisco-sa-cvp-info-dislosure-NZBEwj9V', 'advisoryTitle': 'Cisco Unified Customer Voice Portal Information Disclosure Vulnerability', 'bugIDs': ['CSCvp98656', 'CSCvt45220'], 'ipsSignatures': ['NA'], 'cves': ['CVE-2020-3402'], 'cvrfUrl': 'https://tools.cisco.com/security/center/contentxml/CiscoSecurityAdvisory/cisco-sa-cvp-info-dislosure-NZBEwj9V/cvrf/cisco-sa-cvp-info-dislosure-NZBEwj9V_cvrf.xml', 'cvssBaseScore': '5.3', 'cwe': ['CWE-306'], 'firstPublished': '2020-07-01T16:00:00', 'lastUpdated': '2020-07-01T16:00:00', 'productNames': ['Cisco Unified IP Interactive Voice Response (IVR) '], 'publicationUrl': 'https://tools.cisco.com/security/center/content/CiscoSecurityAdvisory/cisco-sa-cvp-info-dislosure-NZBEwj9V', 'sir': 'Medium', 'summary': '\r\n<p>A vulnerability in the Java Remote Method Invocation (RMI) interface of Cisco&nbsp;Unified Customer Voice Portal (CVP) could allow an unauthenticated, remote attacker to access sensitive information on an affected device.</p>\r\n<p>The vulnerability exists because certain RMI listeners are not properly authenticated. An attacker could exploit this vulnerability by sending a crafted request to the affected listener. A successful exploit could allow the attacker to access sensitive information on an affected device.</p>\r\n<p>There are no workarounds that address this vulnerability.</p>\r\n<p>This advisory is available at the following link:<br><a href="https://tools.cisco.com/security/center/content/CiscoSecurityAdvisory/cisco-sa-cvp-info-dislosure-NZBEwj9V" target="_blank">https://tools.cisco.com/security/center/content/CiscoSecurityAdvisory/cisco-sa-cvp-info-dislosure-NZBEwj9V</a></p>\r\n'}, {'advisoryId': 'cisco-sa-cucm-xss-bLZw4Ctq', 'advisoryTitle': 'Cisco Unified Communications Manager Stored Cross-Site Scripting Vulnerability', 'bugIDs': ['CSCvs88276'], 'ipsSignatures': ['NA'], 'cves': ['CVE-2020-3420'], 'cvrfUrl': 'https://tools.cisco.com/security/center/contentxml/CiscoSecurityAdvisory/cisco-sa-cucm-xss-bLZw4Ctq/cvrf/cisco-sa-cucm-xss-bLZw4Ctq_cvrf.xml', 'cvssBaseScore': '5.4', 'cwe': ['CWE-79'], 'firstPublished': '2020-07-01T16:00:00', 'lastUpdated': '2020-07-01T16:00:00', 'productNames': ['Cisco Unified Communications Manager '], 'publicationUrl': 'https://tools.cisco.com/security/center/content/CiscoSecurityAdvisory/cisco-sa-cucm-xss-bLZw4Ctq', 'sir': 'Medium', 'summary': '\r\n<p>A vulnerability in the web-based management interface of Cisco Unified Communications Manager (Unified CM) and Cisco Unified Communications Manager Session Management Edition (Unified CM SME) could allow an authenticated, remote attacker to conduct a cross-site scripting (XSS) attack against a user of the interface.</p>\r\n<p>The vulnerability is due to insufficient validation of user-supplied input by the web-based management interface of the affected software. An attacker could exploit this vulnerability by inserting malicious data into a specific data field in the interface. A successful exploit could allow the attacker to execute arbitrary script code in the context of the affected interface or access sensitive browser-based information.</p>\r\n<p>There are no workarounds that address this vulnerability.</p>\r\n<p>This advisory is available at the following link:<br><a href="https://tools.cisco.com/security/center/content/CiscoSecurityAdvisory/cisco-sa-cucm-xss-bLZw4Ctq" target="_blank">https://tools.cisco.com/security/center/content/CiscoSecurityAdvisory/cisco-sa-cucm-xss-bLZw4Ctq</a></p>\r\n'}, {'advisoryId': 'cisco-sa-cucm-cuc-imp-xss-OWuSYAp', 'advisoryTitle': 'Cisco Unified Communications Products Cross-Site Scripting Vulnerability', 'bugIDs': ['CSCvs29695', 'CSCvs59653', 'CSCvs59840'], 'ipsSignatures': ['NA'], 'cves': ['CVE-2020-3282'], 'cvrfUrl': 'https://tools.cisco.com/security/center/contentxml/CiscoSecurityAdvisory/cisco-sa-cucm-cuc-imp-xss-OWuSYAp/cvrf/cisco-sa-cucm-cuc-imp-xss-OWuSYAp_cvrf.xml', 'cvssBaseScore': '6.1', 'cwe': ['CWE-79'], 'firstPublished': '2020-07-01T16:00:00', 'lastUpdated': '2020-07-01T16:00:00', 'productNames': ['Cisco Unity Connection ', 'Cisco Unified Communications Manager ', 'Cisco Unified Communications Manager IM and Presence Service ', 'Cisco Unified Communications Manager / Cisco Unity Connection '], 'publicationUrl': 'https://tools.cisco.com/security/center/content/CiscoSecurityAdvisory/cisco-sa-cucm-cuc-imp-xss-OWuSYAp', 'sir': 'Medium', 'summary': '\r\n<p>A vulnerability in the web-based management interface of Cisco Unified Communications Manager, Cisco Unified Communications Manager Session Management Edition, Cisco Unified Communications Manager IM &amp; Presence Service, and Cisco Unity Connection could allow an unauthenticated, remote attacker to conduct a cross-site scripting (XSS) attack against a user of the interface.</p>\r\n<p>The vulnerability is due to insufficient validation of user-supplied input by the web-based management interface of the affected software. An attacker could exploit this vulnerability by persuading a user of the interface to click a crafted link. A successful exploit could allow the attacker to execute arbitrary script code in the context of the affected interface or access sensitive browser-based information.</p>\r\n<p>Cisco has released software updates that address this vulnerability. There are no workarounds that address this vulnerability.</p>\r\n<p>This advisory is available at the following link:<br><a href="https://tools.cisco.com/security/center/content/CiscoSecurityAdvisory/cisco-sa-cucm-cuc-imp-xss-OWuSYAp" target="_blank">https://tools.cisco.com/security/center/content/CiscoSecurityAdvisory/cisco-sa-cucm-cuc-imp-xss-OWuSYAp</a></p>\r\n'}]}

    ciscopsirt(dict)